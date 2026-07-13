#!/usr/bin/env python3
"""cafe_reviews.csv(.xlsx) → 카페 VoC 분석 엑셀 리포트 생성기.

산출물:
  - cafe_reviews.csv            (원본 xlsx에서 추출한 실제 CSV)
  - 카페_VoC_분석리포트.xlsx      (요약 대시보드 + 별점 분포 차트 + 테마 피벗 + 부정 Top3 + 원본)
"""
import csv, io, datetime
from collections import Counter

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import DataBarRule

SRC = "cafe_reviews.csv.xlsx"
CSV_OUT = "cafe_reviews.csv"
XLSX_OUT = "카페_VoC_분석리포트.xlsx"

# ── 팔레트 (커피 톤) ─────────────────────────────────────────
BROWN_DARK = "4E342E"; BROWN = "6F4E37"; GOLD = "C89F65"
CREAM = "F7F3EE"; CREAM_DARK = "EFE5D8"
RED = "C0392B"; GREEN = "2E7D32"; AMBER = "B7791F"
RATING_COLORS = ["C0392B", "E67E22", "F1C40F", "8BC34A", "2E7D32"]  # 1★→5★

thin = Side(style="thin", color="D9CFC3")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
med = Side(style="medium", color=GOLD)
BOX = Border(left=med, right=med, top=med, bottom=med)

F_TITLE = Font(size=16, bold=True, color="FFFFFF")
F_SUB = Font(size=9, color="8D6E63")
F_SECT = Font(size=11, bold=True, color=BROWN_DARK)
F_HEAD = Font(size=10, bold=True, color="FFFFFF")
F_KPI = Font(size=13, bold=True, color=BROWN_DARK)
C = Alignment(horizontal="center", vertical="center")
L = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_range(ws, ref, fill=None, font=None, align=None, border=None):
    for row in ws[ref]:
        for c in row:
            if fill: c.fill = fill
            if font: c.font = font
            if align: c.alignment = align
            if border: c.border = border


def banner(ws, ref, text):
    ws.merge_cells(ref)
    ws[ref.split(":")[0]] = text
    style_range(ws, ref, fill=PatternFill("solid", fgColor=BROWN_DARK), font=F_TITLE, align=C)


def section(ws, ref, text):
    ws.merge_cells(ref)
    ws[ref.split(":")[0]] = text
    style_range(ws, ref, fill=PatternFill("solid", fgColor=CREAM_DARK), font=F_SECT, align=L)


def table_header(ws, row, cols, labels, fill=BROWN):
    for col, label in zip(cols, labels):
        c = ws[f"{col}{row}"]
        c.value = label
        c.fill = PatternFill("solid", fgColor=fill)
        c.font = F_HEAD
        c.alignment = C
        c.border = BORDER


def color_points(chart, colors):
    """단일 시리즈 막대의 점별 색 지정 (미지원 버전이면 단색 유지)."""
    try:
        from openpyxl.chart.series import DataPoint
        pts = []
        for i, col in enumerate(colors):
            dp = DataPoint(idx=i)
            dp.graphicalProperties = GraphicalProperties(solidFill=col)
            pts.append(dp)
        chart.series[0].data_points = pts
    except Exception as e:
        print("  (점별 색 생략:", e, ")")


# ── 1. 원본 로드 (xlsx 안에 CSV 텍스트가 1열로 저장된 형태) ──────
wb_src = openpyxl.load_workbook(SRC, data_only=True)
lines = [r[0] for r in wb_src.active.iter_rows(values_only=True) if r[0]]
rows = list(csv.reader(io.StringIO("\n".join(lines))))
header, data = rows[0], rows[1:]  # 날짜, 플랫폼, 별점, 리뷰, 테마

with open(CSV_OUT, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f); w.writerow(header); w.writerows(data)

# ── 2. 집계 ──────────────────────────────────────────────────
n = len(data)
ratings = [int(r[2]) for r in data]
dist = Counter(ratings)                                   # 별점별 건수
theme_total = Counter(r[4] for r in data)                 # 테마별 전체
theme_by_rating = {t: Counter() for t in theme_total}     # 테마×별점 피벗
for r in data:
    theme_by_rating[r[4]][int(r[2])] += 1
neg_rows = [r for r in data if int(r[2]) <= 2]
pos_rows = [r for r in data if int(r[2]) >= 4]
neg_themes = Counter(r[4] for r in neg_rows).most_common()
pos_themes = Counter(r[4] for r in pos_rows).most_common()
avg = sum(ratings) / n
period = f"{min(r[0] for r in data)} ~ {max(r[0] for r in data)}"
platforms = " / ".join(sorted({r[1] for r in data}))
top_neg, top_neg_cnt = neg_themes[0]
rep_review = {}  # 테마별 대표 리뷰(부정은 최저 별점, 긍정은 최고 별점 우선)
for r in sorted(data, key=lambda r: int(r[2])):
    rep_review.setdefault(("neg", r[4]), r[3])
for r in sorted(data, key=lambda r: -int(r[2])):
    rep_review.setdefault(("pos", r[4]), r[3])

wb = Workbook()

# ── 3. 시트 1 · 📊 요약 대시보드 ─────────────────────────────
ws = wb.active
ws.title = "📊 요약"
ws.sheet_properties.tabColor = BROWN
ws.sheet_view.showGridLines = False
for col, wd in zip("ABCDEFGH", [2, 13, 15, 12, 15, 18, 14, 36]):
    ws.column_dimensions[col].width = wd

banner(ws, "A1:H2", "☕ 카페 VoC(고객의 소리) 분석 리포트")
ws.row_dimensions[1].height = 26; ws.row_dimensions[2].height = 10
ws.merge_cells("A3:H3")
ws["A3"] = (f"분석 기간 {period} · 채널 {platforms} · 리뷰 {n}건 · "
            f"생성일 {datetime.date.today().isoformat()}")
style_range(ws, "A3:H3", font=F_SUB, align=L)

section(ws, "B5:H5", "①  핵심 지표")
kpi_labels = ["총 리뷰 수", "평균 별점", "긍정 리뷰 (4★↑)", "부정 리뷰 (2★↓)", "최다 언급 테마"]
kpi_values = [f"{n}건", f"{avg:.1f}점",
              f"{len(pos_rows)}건 ({len(pos_rows)/n:.0%})",
              f"{len(neg_rows)}건 ({len(neg_rows)/n:.0%})",
              " · ".join(t for t, c in theme_total.most_common() if c == theme_total.most_common(1)[0][1])
              + f" (각 {theme_total.most_common(1)[0][1]}건)"]
table_header(ws, 6, "BCDEF", kpi_labels)
for col, v in zip("BCDEF", kpi_values):
    c = ws[f"{col}7"]; c.value = v; c.font = F_KPI; c.alignment = C; c.border = BORDER
    c.fill = PatternFill("solid", fgColor=CREAM)
ws.row_dimensions[7].height = 24

section(ws, "B9:H9", "②  부정 리뷰 테마 Top 3  (별점 2 이하)")
table_header(ws, 10, "BCDEF", ["순위", "테마", "건수", "부정 내 비중", "대표 리뷰"])
ws.merge_cells("F10:H10")
style_range(ws, "F10:H10", fill=PatternFill("solid", fgColor=BROWN), font=F_HEAD, align=C, border=BORDER)
for i in range(3):
    r = 11 + i
    if i < len(neg_themes):
        t, cnt = neg_themes[i]
        vals = [i + 1, t, f"{cnt}건", cnt / len(neg_rows), f"“{rep_review[('neg', t)]}”"]
    else:
        vals = [i + 1, "—", "—", "—", f"(부정 테마는 총 {len(neg_themes)}종)"]
    for col, v in zip("BCDEF", vals):
        c = ws[f"{col}{r}"]; c.value = v; c.border = BORDER; c.alignment = C
        if isinstance(v, float): c.number_format = "0%"
    ws.merge_cells(f"F{r}:H{r}")
    style_range(ws, f"F{r}:H{r}", align=L, border=BORDER)
    if i == 0:
        style_range(ws, f"B{r}:H{r}", fill=PatternFill("solid", fgColor="FDECEA"))
        for col in "BCDE": ws[f"{col}{r}"].font = Font(bold=True, color=RED)

section(ws, "B15:H15", "③  긍정 리뷰 테마 Top 3  (별점 4 이상)")
table_header(ws, 16, "BCDEF", ["순위", "테마", "건수", "긍정 내 비중", "대표 리뷰"])
ws.merge_cells("F16:H16")
style_range(ws, "F16:H16", fill=PatternFill("solid", fgColor=BROWN), font=F_HEAD, align=C, border=BORDER)
for i in range(3):
    r = 17 + i
    t, cnt = pos_themes[i]
    vals = [i + 1, t, f"{cnt}건", cnt / len(pos_rows), f"“{rep_review[('pos', t)]}”"]
    for col, v in zip("BCDEF", vals):
        c = ws[f"{col}{r}"]; c.value = v; c.border = BORDER; c.alignment = C
        if isinstance(v, float): c.number_format = "0%"
    ws.merge_cells(f"F{r}:H{r}")
    style_range(ws, f"F{r}:H{r}", align=L, border=BORDER)
    if i == 0:
        for col in "BCDE": ws[f"{col}{r}"].font = Font(bold=True, color=GREEN)

section(ws, "B21:H21", "④  핵심 인사이트 & 권고 액션")
pos_names = " · ".join(t for t, _ in pos_themes[:2])
insights = [
    (GREEN, f"🟢 강점: 5점 리뷰 {dist[5]}건의 중심은 '{pos_names}' — 시그니처 메뉴로 SNS·단골 홍보에 적극 활용할 것."),
    (RED, f"🔴 개선 1순위: 부정 리뷰 {len(neg_rows)}건 중 {top_neg_cnt}건({top_neg_cnt/len(neg_rows):.0%})이 "
          f"'{top_neg}' (주말 대기 · 주문 밀림 · 웨이팅) — 재방문을 막는 최대 병목."),
    (AMBER, f"🟡 관찰: '가격' 언급 {theme_total.get('가격', 0)}건(부정 {dict(neg_themes).get('가격', 0)}건 포함)"
            " — 가격 대비 가치(양·세트 구성) 커뮤니케이션 보완 필요."),
]
for i, (color, text) in enumerate(insights):
    r = 22 + i
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = text
    style_range(ws, f"B{r}:H{r}", align=WRAP)
    ws[f"B{r}"].font = Font(size=10, color=color)
    ws.row_dimensions[r].height = 22

ws.merge_cells("B26:H27")
ws["B26"] = (f"📌 결론: 맛·디저트 경쟁력은 검증되었습니다. 지금 재방문의 발목을 잡는 것은 '{top_neg}'입니다."
             " → 피크타임(주말) 인력 보강, 선주문·진동벨 도입 등 웨이팅 개선을 최우선 과제로 권고합니다.")
style_range(ws, "B26:H27", fill=PatternFill("solid", fgColor="FFF8E1"),
            font=Font(size=11, bold=True, color=BROWN_DARK), align=WRAP, border=BOX)

# ── 4. 시트 2 · 별점 분포 (+막대 차트) ───────────────────────
ws2 = wb.create_sheet("별점 분포")
ws2.sheet_properties.tabColor = GOLD
ws2.sheet_view.showGridLines = False
for col, wd in zip("ABC", [10, 10, 10]):
    ws2.column_dimensions[col].width = wd
banner(ws2, "A1:C1", "⭐ 별점 분포")
ws2.row_dimensions[1].height = 24
table_header(ws2, 3, "ABC", ["별점", "건수", "비율"])
for i in range(1, 6):
    r = 3 + i
    ws2[f"A{r}"] = f"{i}★"; ws2[f"B{r}"] = dist.get(i, 0)
    ws2[f"C{r}"] = dist.get(i, 0) / n; ws2[f"C{r}"].number_format = "0%"
    for col in "ABC":
        ws2[f"{col}{r}"].border = BORDER; ws2[f"{col}{r}"].alignment = C
ws2["A9"] = "합계"; ws2["B9"] = n; ws2["C9"] = 1; ws2["C9"].number_format = "0%"
style_range(ws2, "A9:C9", fill=PatternFill("solid", fgColor=CREAM_DARK),
            font=Font(bold=True), align=C, border=BORDER)

ch = BarChart(); ch.type = "col"; ch.title = "별점 분포 (건수)"
ch.add_data(Reference(ws2, min_col=2, min_row=3, max_row=8), titles_from_data=True)
ch.set_categories(Reference(ws2, min_col=1, min_row=4, max_row=8))
ch.legend = None; ch.gapWidth = 60; ch.width = 15; ch.height = 9
ch.dataLabels = DataLabelList(); ch.dataLabels.showVal = True
ch.series[0].graphicalProperties = GraphicalProperties(solidFill=BROWN)
color_points(ch, RATING_COLORS)
ws2.add_chart(ch, "E3")

# ── 5. 시트 3 · 테마별 피벗 ──────────────────────────────────
ws3 = wb.create_sheet("테마별 피벗")
ws3.sheet_properties.tabColor = "A1887F"
ws3.sheet_view.showGridLines = False
for col, wd in zip("ABCDEFGHI", [12, 7, 7, 7, 7, 7, 9, 10, 10]):
    ws3.column_dimensions[col].width = wd
banner(ws3, "A1:I1", "🧩 테마별 리뷰 수 피벗 (테마 × 별점)")
ws3.row_dimensions[1].height = 24
table_header(ws3, 3, "ABCDEFGHI",
             ["테마", "1★", "2★", "3★", "4★", "5★", "합계", "부정(≤2)", "긍정(≥4)"])
themes_sorted = sorted(theme_total, key=lambda t: (-theme_total[t], t))
for i, t in enumerate(themes_sorted):
    r = 4 + i
    by = theme_by_rating[t]
    neg_c = by[1] + by[2]; pos_c = by[4] + by[5]
    for col, v in zip("ABCDEFGHI",
                      [t, by[1], by[2], by[3], by[4], by[5], theme_total[t], neg_c, pos_c]):
        c = ws3[f"{col}{r}"]; c.value = v; c.border = BORDER; c.alignment = C
    ws3[f"A{r}"].font = Font(bold=True)
    if neg_c: ws3[f"H{r}"].font = Font(bold=True, color=RED)
    if pos_c: ws3[f"I{r}"].font = Font(color=GREEN)
last = 3 + len(themes_sorted)
tot_r = last + 1
totals = ["전체", dist[1], dist[2], dist[3], dist[4], dist[5], n, len(neg_rows), len(pos_rows)]
for col, v in zip("ABCDEFGHI", totals):
    ws3[f"{col}{tot_r}"] = v
style_range(ws3, f"A{tot_r}:I{tot_r}", fill=PatternFill("solid", fgColor=CREAM_DARK),
            font=Font(bold=True), align=C, border=BORDER)
ws3.conditional_formatting.add(
    f"G4:G{last}",
    DataBarRule(start_type="num", start_value=0, end_type="num",
                end_value=max(theme_total.values()), color=GOLD, showValue=True))

ch3 = BarChart(); ch3.type = "bar"; ch3.title = "테마별 언급량"
ch3.add_data(Reference(ws3, min_col=7, min_row=3, max_row=last), titles_from_data=True)
ch3.set_categories(Reference(ws3, min_col=1, min_row=4, max_row=last))
ch3.legend = None; ch3.width = 13; ch3.height = 9
ch3.dataLabels = DataLabelList(); ch3.dataLabels.showVal = True
ch3.series[0].graphicalProperties = GraphicalProperties(solidFill=BROWN)
ws3.add_chart(ch3, "K3")

# ── 6. 시트 4 · 부정 리뷰 (Top3 + 목록 + 차트) ───────────────
ws4 = wb.create_sheet("부정 리뷰")
ws4.sheet_properties.tabColor = RED
ws4.sheet_view.showGridLines = False
for col, wd in zip("ABCDE", [12, 10, 8, 46, 12]):
    ws4.column_dimensions[col].width = wd
banner(ws4, "A1:E1", f"🚨 부정 리뷰 분석 (별점 2 이하 · {len(neg_rows)}건)")
ws4.row_dimensions[1].height = 24
table_header(ws4, 3, "ABCD", ["순위", "테마", "건수", "부정 내 비중"], fill=RED)
for i in range(3):
    r = 4 + i
    if i < len(neg_themes):
        t, cnt = neg_themes[i]
        vals = [i + 1, t, cnt, cnt / len(neg_rows)]
    else:
        vals = [i + 1, "—", "—", "—"]
    for col, v in zip("ABCD", vals):
        c = ws4[f"{col}{r}"]; c.value = v; c.border = BORDER; c.alignment = C
        if isinstance(v, float): c.number_format = "0%"
    if i == 0:
        style_range(ws4, f"A{r}:D{r}", fill=PatternFill("solid", fgColor="FDECEA"))
        for col in "ABCD": ws4[f"{col}{r}"].font = Font(bold=True, color=RED)

section(ws4, "A8:E8", "부정 리뷰 전체 목록")
table_header(ws4, 9, "ABCDE", header, fill=RED)
for i, row in enumerate(neg_rows):
    r = 10 + i
    for col, v in zip("ABCDE", [row[0], row[1], int(row[2]), row[3], row[4]]):
        c = ws4[f"{col}{r}"]; c.value = v; c.border = BORDER
        c.alignment = L if col == "D" else C
    ws4[f"C{r}"].font = Font(bold=True, color=RED)

ch4 = BarChart(); ch4.type = "col"; ch4.title = "부정 테마 분포"
ch4.add_data(Reference(ws4, min_col=3, min_row=3, max_row=3 + len(neg_themes)), titles_from_data=True)
ch4.set_categories(Reference(ws4, min_col=2, min_row=4, max_row=3 + len(neg_themes)))
ch4.legend = None; ch4.gapWidth = 60; ch4.width = 11; ch4.height = 7
ch4.dataLabels = DataLabelList(); ch4.dataLabels.showVal = True
ch4.series[0].graphicalProperties = GraphicalProperties(solidFill=RED)
ws4.add_chart(ch4, "G3")

# ── 7. 시트 5 · 원본 데이터 ──────────────────────────────────
ws5 = wb.create_sheet("원본 데이터")
ws5.sheet_properties.tabColor = "9E9E9E"
for col, wd in zip("ABCDE", [12, 10, 8, 48, 12]):
    ws5.column_dimensions[col].width = wd
table_header(ws5, 1, "ABCDE", header)
for i, row in enumerate(data):
    r = 2 + i
    rating = int(row[2])
    for col, v in zip("ABCDE", [row[0], row[1], rating, row[3], row[4]]):
        c = ws5[f"{col}{r}"]; c.value = v; c.border = BORDER
        c.alignment = L if col == "D" else C
        if i % 2: c.fill = PatternFill("solid", fgColor=CREAM)
    ws5[f"C{r}"].font = Font(bold=True,
                             color=RED if rating <= 2 else (GREEN if rating >= 4 else AMBER))
ws5.freeze_panes = "A2"
ws5.auto_filter.ref = f"A1:E{1 + n}"

wb.save(XLSX_OUT)
print(f"저장 완료 → {XLSX_OUT}")
print(f"  리뷰 {n}건 · 평균 {avg:.2f}점 · 부정 {len(neg_rows)}건")
print(f"  부정 Top3: {neg_themes[:3]}")
print(f"  긍정 Top3: {pos_themes[:3]}")
